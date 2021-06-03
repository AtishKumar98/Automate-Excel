from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data =  {
    "Joe":{
        
        "maths":65,
        "science":78,
        "english":98,
        "Gym":40},

    "Bill":{
        
        "maths":35,
        "science":74,
        "english":49,
        "Gym":50},
    "Tim":{
        
        "maths":45,
        "science":58,
        "english":78,
        "Gym":69},
    "Sally":{
        
        "maths":75,
        "science":58,
        "english":88,
        "Gym":79},
    "Jane":{
        
        "maths":75,
        "science":39,
        "english":90,
        "Gym":79},
}

wb = Workbook()
ws = wb.active
ws.title ="New Data"
headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)
for col in range(2,len(data["Joe"])+2):
    char = get_column_letter(col)
    ws[char+"7"] = f"=SUM({char+'2'}:{char+'6'})/{len(data)}"

wb.save("NewAutom.xlsx")

